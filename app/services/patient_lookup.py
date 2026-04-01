"""
Structured Patient Lookup
=========================
Deterministic row-level lookup for patient queries over uploaded CSV/XLSX files.

Why this exists:
- Patient prompts are naturally structured (gender, city, outcome, doctor).
- Generic semantic retrieval over chunked tabular text can miss exact row matches.
- This module scans the raw tabular uploads directly and returns grounded,
  citation-friendly row summaries for investor-demo reliability.
"""

from __future__ import annotations

import csv
import logging
import os
import re
import time
from dataclasses import dataclass, field
from functools import lru_cache
from typing import Optional

from langchain_core.documents import Document

from models.response import ChunkSource

logger = logging.getLogger(__name__)

_MAX_RESULTS = 8
_SUPPORTED_EXTENSIONS = {
    ".csv": "csv",
    ".xlsx": "excel",
}
_HEADER_ALIASES = {
    "patient_id": "patient_id",
    "patientid": "patient_id",
    "name": "name",
    "age": "age",
    "gender": "gender",
    "issue_type": "issue_type",
    "issuetype": "issue_type",
    "doctor_name": "doctor_name",
    "doctorname": "doctor_name",
    "doctor": "doctor_name",
    "department": "department",
    "admission_date": "admission_date",
    "admissiondate": "admission_date",
    "discharge_date": "discharge_date",
    "dischargedate": "discharge_date",
    "city": "city",
    "country": "country",
    "insurance_provider": "insurance_provider",
    "insuranceprovider": "insurance_provider",
    "severity": "severity",
    "outcome": "outcome",
}
_PATIENT_TABLE_FIELDS = {
    "patient_id",
    "name",
    "gender",
    "doctor_name",
    "city",
    "outcome",
    "issue_type",
    "department",
}
_DISPLAY_FIELDS = [
    ("Patient_ID", "patient_id"),
    ("Name", "name"),
    ("Age", "age"),
    ("Gender", "gender"),
    ("Issue_Type", "issue_type"),
    ("Doctor_Name", "doctor_name"),
    ("Department", "department"),
    ("Admission_Date", "admission_date"),
    ("Discharge_Date", "discharge_date"),
    ("City", "city"),
    ("Country", "country"),
    ("Insurance_Provider", "insurance_provider"),
    ("Severity", "severity"),
    ("Outcome", "outcome"),
]
_SPACE_PATTERN = re.compile(r"\s+")
_NON_ALNUM_PATTERN = re.compile(r"[^a-z0-9]+")
_GENDER_PATTERN = re.compile(r"(?i)\b(male|female)\b")
_TITLE_PATTERN = re.compile(r"(?i)\bdr\.?\s*")


@dataclass(slots=True)
class PatientQueryFilters:
    gender: Optional[str] = None
    city: Optional[str] = None
    outcome: Optional[str] = None
    doctor_name: Optional[str] = None

    def as_log_dict(self) -> dict[str, str]:
        return {
            key: value
            for key, value in {
                "gender": self.gender,
                "city": self.city,
                "outcome": self.outcome,
                "doctor_name": self.doctor_name,
            }.items()
            if value
        }

    def has_any(self) -> bool:
        return any((self.gender, self.city, self.outcome, self.doctor_name))

    def active_filter_count(self) -> int:
        return len(self.as_log_dict())


@dataclass(slots=True)
class PatientTableRow:
    source: str
    doc_type: str
    sheet_name: str
    row_number: int
    user_id: Optional[str]
    fields: dict[str, str] = field(default_factory=dict)

    @property
    def page_or_sheet(self) -> str:
        return f"sheet_{self.sheet_name}_row_{self.row_number}"

    @property
    def chunk_id(self) -> str:
        return f"structured:{self.source}:{self.sheet_name}:{self.row_number}"

    def get(self, key: str) -> str:
        return self.fields.get(key, "")


@dataclass(slots=True)
class StructuredPatientLookupResult:
    handled: bool
    answer: str
    sources: list[ChunkSource]
    documents: list[Document]
    matched_rows: int
    returned_rows: int
    scanned_rows: int
    scanned_files: int
    filters: PatientQueryFilters
    latency_ms: float


@lru_cache(maxsize=64)
def _load_csv_rows_cached(file_path: str, mtime: float, user_id: Optional[str]) -> tuple[PatientTableRow, ...]:
    with open(file_path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = [_canonical_header(header or "") for header in (reader.fieldnames or [])]
        if not _looks_like_patient_table(headers):
            return ()

        rows: list[PatientTableRow] = []
        for row_index, raw_row in enumerate(reader, start=1):
            fields = _normalise_row(raw_row)
            if not any(fields.values()):
                continue
            rows.append(
                PatientTableRow(
                    source=os.path.basename(file_path),
                    doc_type="csv",
                    sheet_name="Sheet1",
                    row_number=row_index,
                    user_id=user_id,
                    fields=fields,
                )
            )
        return tuple(rows)


@lru_cache(maxsize=32)
def _load_excel_rows_cached(file_path: str, mtime: float, user_id: Optional[str]) -> tuple[PatientTableRow, ...]:
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    rows: list[PatientTableRow] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        values = list(worksheet.iter_rows(values_only=True))
        if not values:
            continue

        headers = [_canonical_header(str(value or "")) for value in values[0]]
        if not _looks_like_patient_table(headers):
            continue

        for row_index, values_row in enumerate(values[1:], start=1):
            raw_row = {
                str(values[0][index] or f"col_{index}"): values_row[index] if index < len(values_row) else ""
                for index in range(len(values[0]))
            }
            fields = _normalise_row(raw_row)
            if not any(fields.values()):
                continue
            rows.append(
                PatientTableRow(
                    source=os.path.basename(file_path),
                    doc_type="excel",
                    sheet_name=sheet_name,
                    row_number=row_index,
                    user_id=user_id,
                    fields=fields,
                )
            )

    return tuple(rows)


def lookup_patient_rows(
    question: str,
    *,
    upload_dir: str,
    user_id: Optional[str],
    source_filter: Optional[str] = None,
    doc_type_filter: Optional[str] = None,
    max_results: int = _MAX_RESULTS,
) -> StructuredPatientLookupResult:
    started = time.perf_counter()
    question_normalized = _normalize_text(question)

    if not question_normalized:
        return StructuredPatientLookupResult(
            handled=False,
            answer="",
            sources=[],
            documents=[],
            matched_rows=0,
            returned_rows=0,
            scanned_rows=0,
            scanned_files=0,
            filters=PatientQueryFilters(),
            latency_ms=0.0,
        )

    candidate_files = list(_iter_candidate_files(upload_dir, user_id, source_filter, doc_type_filter))
    if not candidate_files:
        return StructuredPatientLookupResult(
            handled=False,
            answer="",
            sources=[],
            documents=[],
            matched_rows=0,
            returned_rows=0,
            scanned_rows=0,
            scanned_files=0,
            filters=PatientQueryFilters(),
            latency_ms=round((time.perf_counter() - started) * 1000, 1),
        )

    rows: list[PatientTableRow] = []
    for file_path, doc_type in candidate_files:
        rows.extend(_load_rows(file_path, doc_type, user_id))

    if not rows:
        return StructuredPatientLookupResult(
            handled=False,
            answer="",
            sources=[],
            documents=[],
            matched_rows=0,
            returned_rows=0,
            scanned_rows=0,
            scanned_files=len(candidate_files),
            filters=PatientQueryFilters(),
            latency_ms=round((time.perf_counter() - started) * 1000, 1),
        )

    filters = _derive_filters(question_normalized, rows)
    if not filters.has_any():
        return StructuredPatientLookupResult(
            handled=False,
            answer="",
            sources=[],
            documents=[],
            matched_rows=0,
            returned_rows=0,
            scanned_rows=len(rows),
            scanned_files=len(candidate_files),
            filters=filters,
            latency_ms=round((time.perf_counter() - started) * 1000, 1),
        )

    matched_rows = [row for row in rows if _row_matches(row, filters)]
    visible_rows = matched_rows[:max(1, max_results)]
    sources = [_row_to_source(row) for row in visible_rows]
    documents = [_row_to_document(row) for row in visible_rows]
    answer = _build_answer(filters, matched_rows, visible_rows)
    latency_ms = round((time.perf_counter() - started) * 1000, 1)

    logger.info(
        "Structured patient lookup | user=%s | files=%d | scanned_rows=%d | matched_rows=%d | returned_rows=%d | filters=%s",
        user_id,
        len(candidate_files),
        len(rows),
        len(matched_rows),
        len(visible_rows),
        filters.as_log_dict(),
    )

    return StructuredPatientLookupResult(
        handled=True,
        answer=answer,
        sources=sources,
        documents=documents,
        matched_rows=len(matched_rows),
        returned_rows=len(visible_rows),
        scanned_rows=len(rows),
        scanned_files=len(candidate_files),
        filters=filters,
        latency_ms=latency_ms,
    )


def _iter_candidate_files(
    upload_dir: str,
    user_id: Optional[str],
    source_filter: Optional[str],
    doc_type_filter: Optional[str],
) -> list[tuple[str, str]]:
    if doc_type_filter and doc_type_filter not in {"csv", "excel"}:
        return []

    roots: list[str] = []
    if user_id:
        roots.append(os.path.join(upload_dir, user_id))
    roots.append(upload_dir)

    discovered: list[tuple[str, str]] = []
    seen_paths: set[str] = set()
    normalized_source_filter = _normalize_text(source_filter or "")

    for root in roots:
        if not os.path.isdir(root):
            continue
        for entry in sorted(os.listdir(root)):
            path = os.path.join(root, entry)
            if path in seen_paths or not os.path.isfile(path):
                continue
            ext = os.path.splitext(entry)[1].lower()
            doc_type = _SUPPORTED_EXTENSIONS.get(ext)
            if not doc_type:
                continue
            if doc_type_filter and doc_type != doc_type_filter:
                continue
            if normalized_source_filter and normalized_source_filter not in _normalize_text(entry):
                continue
            seen_paths.add(path)
            discovered.append((path, doc_type))

    return discovered


def _load_rows(file_path: str, doc_type: str, user_id: Optional[str]) -> list[PatientTableRow]:
    mtime = os.path.getmtime(file_path)
    if doc_type == "csv":
        return list(_load_csv_rows_cached(file_path, mtime, user_id))
    return list(_load_excel_rows_cached(file_path, mtime, user_id))


def _derive_filters(question_normalized: str, rows: list[PatientTableRow]) -> PatientQueryFilters:
    gender_match = _GENDER_PATTERN.search(question_normalized)
    gender = gender_match.group(1).lower() if gender_match else None

    city = _match_inventory_value(question_normalized, [row.get("city") for row in rows])
    outcome = _match_inventory_value(question_normalized, [row.get("outcome") for row in rows])
    doctor_name = _match_inventory_value(
        _normalize_doctor(question_normalized),
        [_normalize_doctor(row.get("doctor_name")) for row in rows],
        normalizer=_normalize_doctor,
    )

    if doctor_name:
        for row in rows:
            if _normalize_doctor(row.get("doctor_name")) == _normalize_doctor(doctor_name):
                doctor_name = row.get("doctor_name")
                break

    return PatientQueryFilters(
        gender=gender.title() if gender else None,
        city=city,
        outcome=outcome,
        doctor_name=doctor_name,
    )


def _match_inventory_value(
    question_normalized: str,
    values: list[str],
    *,
    normalizer=None,
) -> Optional[str]:
    seen: dict[str, str] = {}
    for value in values:
        cleaned = (value or "").strip()
        if not cleaned:
            continue
        normalized_value = normalizer(cleaned) if normalizer else _normalize_text(cleaned)
        if not normalized_value:
            continue
        seen.setdefault(normalized_value, cleaned)

    for normalized_value in sorted(seen, key=len, reverse=True):
        if re.search(rf"\b{re.escape(normalized_value)}\b", question_normalized):
            return seen[normalized_value]
    return None


def _row_matches(row: PatientTableRow, filters: PatientQueryFilters) -> bool:
    if filters.gender and _normalize_text(row.get("gender")) != _normalize_text(filters.gender):
        return False
    if filters.city and _normalize_text(row.get("city")) != _normalize_text(filters.city):
        return False
    if filters.outcome and _normalize_text(row.get("outcome")) != _normalize_text(filters.outcome):
        return False
    if filters.doctor_name and _normalize_doctor(row.get("doctor_name")) != _normalize_doctor(filters.doctor_name):
        return False
    return True


def _row_to_source(row: PatientTableRow) -> ChunkSource:
    return ChunkSource(
        source=row.source,
        page_or_sheet=row.page_or_sheet,
        doc_type=row.doc_type,
        chunk_id=row.chunk_id,
        user_id=row.user_id,
        snippet=_render_row_summary(row),
    )


def _row_to_document(row: PatientTableRow) -> Document:
    return Document(
        page_content=_render_row_summary(row),
        metadata={
            "source": row.source,
            "page_or_sheet": row.page_or_sheet,
            "doc_type": row.doc_type,
            "chunk_id": row.chunk_id,
            "user_id": row.user_id,
        },
    )


def _build_answer(
    filters: PatientQueryFilters,
    matched_rows: list[PatientTableRow],
    visible_rows: list[PatientTableRow],
) -> str:
    if not matched_rows:
        return "This information is not available in the provided documents."

    qualifier_parts: list[str] = []
    if filters.gender:
        qualifier_parts.append(filters.gender.lower())
    qualifier_parts.append("patient" if len(matched_rows) == 1 else "patients")
    if filters.city:
        qualifier_parts.append(f"in {filters.city}")
    if filters.outcome:
        qualifier_parts.append(f"with outcome {filters.outcome}")
    if filters.doctor_name:
        qualifier_parts.append(f"for {filters.doctor_name}")

    intro = f"Found {len(matched_rows)} matching {' '.join(qualifier_parts).strip()}."
    if len(visible_rows) < len(matched_rows):
        intro += f" Showing first {len(visible_rows)} results."

    lines = [intro]
    for row in visible_rows:
        lines.append(
            f"- {_render_row_summary(row)} [source: {row.source}, {row.page_or_sheet}]"
        )
    return "\n".join(lines)


def _render_row_summary(row: PatientTableRow) -> str:
    parts = [
        f"{label}: {value}"
        for label, key in _DISPLAY_FIELDS
        if (value := row.get(key))
    ]
    return " | ".join(parts)


def _normalise_row(raw_row: dict[object, object]) -> dict[str, str]:
    fields: dict[str, str] = {}
    for raw_header, raw_value in raw_row.items():
        header = _canonical_header(str(raw_header or ""))
        if not header:
            continue
        value = _SPACE_PATTERN.sub(" ", str(raw_value or "").strip())
        fields[header] = value
    return fields


def _canonical_header(header: str) -> str:
    cleaned = _NON_ALNUM_PATTERN.sub("_", header.strip().lower()).strip("_")
    return _HEADER_ALIASES.get(cleaned, cleaned)


def _looks_like_patient_table(headers: list[str]) -> bool:
    header_set = {header for header in headers if header}
    matched = header_set & _PATIENT_TABLE_FIELDS
    return len(matched) >= 4 and "name" in matched and "gender" in matched


def _normalize_text(value: str) -> str:
    return _SPACE_PATTERN.sub(" ", value.strip().lower())


def _normalize_doctor(value: str) -> str:
    without_title = _TITLE_PATTERN.sub("", value or "")
    return _normalize_text(without_title)

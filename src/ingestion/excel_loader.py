"""
Excel / CSV Loader — Table-Aware
==================================
Strategy:
  - Excel (.xlsx): reads each sheet, converts rows to readable text blocks.
  - CSV:           reads the file as a single sheet.

Each sheet becomes one or more Documents.
Rows are serialised as "Column: Value" pairs for embedding quality.

Metadata per sheet:
  source, sheet_name, row_range, doc_type="excel"/"csv", user_id
"""

import csv
import logging
import os
from typing import List

from langchain_core.documents import Document

logger = logging.getLogger(__name__)

ROWS_PER_CHUNK = 50       # rows per Document to control chunk size
MAX_CELL_CHARS = 500      # truncate runaway cell values (XML/JSON blobs in CSVs)
                          # prevents one CSV from exploding into 28k+ chunks


def load_excel(file_path: str, user_id: str | None = None) -> List[Document]:
    """Load an Excel file — each sheet becomes chunked Documents."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel not found: {file_path}")

    filename = os.path.basename(file_path)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        documents: List[Document] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]

            docs = _rows_to_documents(
                headers=headers,
                data_rows=data_rows,
                filename=filename,
                sheet_name=sheet_name,
                doc_type="excel",
                user_id=user_id,
            )
            documents.extend(docs)

        logger.info("Excel loaded | sheets=%d | docs=%d | file=%s",
                    len(wb.sheetnames), len(documents), filename)
        return documents

    except Exception as e:
        logger.exception("Failed to load Excel: %s", file_path)
        raise RuntimeError(f"Excel load failed: {filename}") from e


def load_csv(file_path: str, user_id: str | None = None) -> List[Document]:
    """Load a CSV file — returns chunked Documents."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"CSV not found: {file_path}")

    filename = os.path.basename(file_path)

    try:
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            return []

        headers = rows[0]
        data_rows = [tuple(r) for r in rows[1:]]

        documents = _rows_to_documents(
            headers=headers,
            data_rows=data_rows,
            filename=filename,
            sheet_name="Sheet1",
            doc_type="csv",
            user_id=user_id,
        )

        logger.info("CSV loaded | rows=%d | docs=%d | file=%s",
                    len(data_rows), len(documents), filename)
        return documents

    except Exception as e:
        logger.exception("Failed to load CSV: %s", file_path)
        raise RuntimeError(f"CSV load failed: {filename}") from e


def _rows_to_documents(
    headers: List[str],
    data_rows: list,
    filename: str,
    sheet_name: str,
    doc_type: str,
    user_id: str | None,
) -> List[Document]:
    """Convert tabular rows into chunked LangChain Documents."""
    documents: List[Document] = []

    for chunk_start in range(0, len(data_rows), ROWS_PER_CHUNK):
        chunk_rows = data_rows[chunk_start: chunk_start + ROWS_PER_CHUNK]
        chunk_end = chunk_start + len(chunk_rows)

        lines = []
        for row in chunk_rows:
            parts = []
            for i, cell in enumerate(row):
                if i >= len(headers):
                    break
                value = str(cell) if cell is not None else ""
                # Truncate long blobs — preserves key info, kills chunk explosion
                if len(value) > MAX_CELL_CHARS:
                    value = value[:MAX_CELL_CHARS] + "…[truncated]"
                parts.append(f"{headers[i]}: {value}")
            lines.append(" | ".join(parts))

        text = "\n".join(lines)
        if not text.strip():
            continue

        documents.append(Document(
            page_content=text,
            metadata={
                "source": filename,
                "sheet_name": sheet_name,
                "page_or_sheet": f"sheet_{sheet_name}_rows_{chunk_start + 1}_{chunk_end}",
                "row_range": f"{chunk_start + 1}-{chunk_end}",
                "doc_type": doc_type,
                "user_id": user_id,
            }
        ))

    return documents
